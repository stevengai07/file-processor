# -*- coding: utf-8 -*-
"""Excel merge helper functions for similar-template file merging."""

from __future__ import annotations

import argparse
import io
import os
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter


def normalize_header(header: Any) -> str:
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = re.sub(r"[\s\-_.()（）\[\]{}:：;；,，\"']+", " ", text)
    text = re.sub(r"[^\w\u4e00-\u9fff ]+", "", text)
    return " ".join(text.split())


def header_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    a_norm = normalize_header(a)
    b_norm = normalize_header(b)
    if a_norm == b_norm:
        return 1.0
    return SequenceMatcher(None, a_norm, b_norm).ratio()


def jaccard_similarity(a: set, b: set) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    intersection = a & b
    union = a | b
    return len(intersection) / len(union)


def read_excel_sheet_names(file_bytes: bytes, filename: str) -> List[str]:
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        return list(xls.sheet_names)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件 {filename}：{exc}") from exc



# ----- Smart Excel reading helpers to handle multi-row headers -----

def _find_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 5) -> int:
    """Find the actual header row index in a multi-row header Excel file.

    Heuristic: detect a title row + numeric column index row pattern, or scan
    for the first row where the majority of non-empty cells contain non-numeric text.
    """
    def is_numeric_row(series: pd.Series) -> bool:
        non_empty = series.dropna()
        if len(non_empty) < 2:
            return False
        return all(
            isinstance(v, (int, float)) or (isinstance(v, str) and str(v).strip().isdigit())
            for v in non_empty
        )

    def is_text_row(series: pd.Series) -> bool:
        non_empty = series.dropna()
        if len(non_empty) < 1:
            return False
        text_cells = 0
        for v in non_empty:
            if isinstance(v, str):
                s = v.strip()
                if s and not s.isdigit():
                    text_cells += 1
        return text_cells >= max(1, len(non_empty) * 0.6)

    if len(df_raw) >= 3:
        row0, row1, row2 = df_raw.iloc[0], df_raw.iloc[1], df_raw.iloc[2]
        if is_text_row(row0) and is_numeric_row(row1) and is_text_row(row2):
            return 2

    scan_rows = min(max_scan_rows, len(df_raw))
    for i in range(scan_rows):
        row = df_raw.iloc[i]
        non_empty = row.dropna()
        if len(non_empty) >= 3:
            text_cells = 0
            for v in non_empty:
                if isinstance(v, str):
                    s = v.strip()
                    if s and not s.isdigit():
                        text_cells += 1
            if text_cells >= len(non_empty) * 0.6:
                return i
    return 0


def _read_excel_smart(file_bytes: bytes, sheet_name: str = 0) -> pd.DataFrame:
    """Read Excel file with automatic header row detection.

    Returns a DataFrame with proper header (column names) and rows starting at
    the first data row after the header.
    """
    try:
        # Read the first N rows without header to detect header row
        df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, sheet_name=sheet_name, nrows=10)
    except Exception as exc:
        # Fallback to normal read to surface the original error
        raise ValueError(f"无法读取 Excel 文件（用于检测表头）：{exc}") from exc

    header_row = _find_header_row(df_raw, max_scan_rows=min(5, len(df_raw)))

    # Now read with the detected header row
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), header=header_row, sheet_name=sheet_name, dtype=str)
    except Exception as exc:
        raise ValueError(f"读取工作表数据失败（按检测到的表头）：{exc}") from exc

    # Drop completely empty rows
    df = df.dropna(how="all")
    # Drop rows where the first column is NaN (likely leftover header artifacts)
    if df.shape[1] > 0:
        df = df[df.iloc[:, 0].notna()]
    # Ensure string column names
    df.columns = [str(c) for c in df.columns]
    # Reset index
    df = df.reset_index(drop=True)
    return df


def read_excel_sheet_columns(file_bytes: bytes, sheet_name: str, filename: str) -> List[str]:
    try:
        df = _read_excel_smart(file_bytes, sheet_name=sheet_name)
        df = df.dropna(axis=1, how="all")
        df.columns = [str(c) for c in df.columns]
        return list(df.columns)
    except Exception as exc:
        raise ValueError(f"无法读取工作表 '{sheet_name}' 的列：{exc}") from exc


def read_excel_sheet_dataframe(file_bytes: bytes, sheet_name: str, filename: str) -> pd.DataFrame:
    try:
        df = _read_excel_smart(file_bytes, sheet_name=sheet_name)
        # 保证首列有意义数据，避免把标题行或空行当成数据
        if df.shape[1] > 1:
            df = df[df.iloc[:, 0].notna()]
        return df.reset_index(drop=True)
    except Exception as exc:
        raise ValueError(f"读取工作表 '{sheet_name}' 数据失败：{exc}") from exc


def compute_similarity_score(headers_list: List[List[str]]) -> float:
    if len(headers_list) < 2:
        return 1.0
    total = 0.0
    count = 0
    normalized_sets = [set(normalize_header(h) for h in headers) for headers in headers_list]
    for i in range(len(normalized_sets)):
        for j in range(i + 1, len(normalized_sets)):
            total += jaccard_similarity(normalized_sets[i], normalized_sets[j])
            count += 1
    return total / count if count else 1.0


def all_headers_exact_match(headers_list: List[List[str]]) -> bool:
    if not headers_list:
        return True
    normalized = [set(normalize_header(h) for h in headers) for headers in headers_list]
    first = normalized[0]
    return all(hs == first for hs in normalized[1:])


def build_canonical_headers(headers_list: List[List[str]], similarity_threshold: float = 0.75) -> List[str]:
    """Build unified output headers using local aliases and similarity rules."""
    canonical: List[str] = []
    canonical_norm: List[str] = []

    alias_targets = {
        "姓名": {"姓名", "有效证件姓名", "申请人姓名", "客户姓名", "人员姓名", "用户姓名"},
        "联系电话": {"联系电话", "联系电话号码", "手机", "手机号", "手机号码", "电话", "电话号码"},
        "出生日期": {"出生日期", "出生年月", "生日", "出生时间", "出生日期年月"},
        "身份证号": {"身份证号", "身份证号码", "证件号", "证件号码", "有效证件号码"},
        "地址": {"地址", "联系地址", "家庭住址", "居住地址", "开户地址"},
        "公司名称": {"公司名称", "单位名称", "企业名称", "机构名称", "所属公司"},
        "电子邮箱": {"电子邮箱", "邮箱", "email", "e mail"},
    }
    alias_targets = {
        normalize_header(target): {normalize_header(value) for value in values}
        for target, values in alias_targets.items()
    }

    for headers in headers_list:
        for header in headers:
            if not header:
                continue

            raw_header = str(header)
            raw_norm = normalize_header(raw_header)
            if not raw_norm:
                continue

            output_header = raw_header
            for target, aliases in alias_targets.items():
                if raw_norm in aliases:
                    output_header = target
                    break

            output_norm = normalize_header(output_header)
            best_match = 0.0
            best_idx: Optional[int] = None

            for idx, existing in enumerate(canonical_norm):
                ratio = header_similarity(output_norm, existing)
                if ratio > best_match:
                    best_match = ratio
                    best_idx = idx

            if best_idx is not None and best_match >= similarity_threshold:
                continue

            canonical.append(output_header)
            canonical_norm.append(output_norm)

    return canonical


def suggest_column_mapping(file_columns: List[str], canonical_headers: List[str]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    canonical_norm = [normalize_header(h) for h in canonical_headers]

    alias_groups = [
        {"姓名", "有效证件姓名", "申请人姓名", "客户姓名", "人员姓名", "用户姓名"},
        {"联系电话", "联系电话号码", "手机", "手机号", "手机号码", "电话", "电话号码"},
        {"出生日期", "出生年月", "生日", "出生时间", "出生日期年月"},
        {"身份证号", "身份证号码", "证件号", "证件号码", "有效证件号码"},
        {"地址", "联系地址", "家庭住址", "居住地址", "开户地址"},
        {"公司名称", "单位名称", "企业名称", "机构名称", "所属公司"},
        {"电子邮箱", "邮箱", "email", "e mail"},
    ]
    alias_groups = [
        {normalize_header(value) for value in group}
        for group in alias_groups
    ]

    for column in file_columns:
        col_norm = normalize_header(column)

        alias_target = None
        for group in alias_groups:
            if col_norm in group:
                for index, target_norm in enumerate(canonical_norm):
                    if target_norm in group:
                        alias_target = canonical_headers[index]
                        break
            if alias_target:
                break

        if alias_target:
            result[column] = alias_target
            continue

        if col_norm in canonical_norm:
            result[column] = canonical_headers[canonical_norm.index(col_norm)]
            continue
        best_score = 0.0
        best_idx: Optional[int] = None
        for idx, candidate_norm in enumerate(canonical_norm):
            score = header_similarity(col_norm, candidate_norm)
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_idx is not None and best_score >= 0.6:
            result[column] = canonical_headers[best_idx]
        else:
            result[column] = column
    return result


def normalize_mapping(mapping: Dict[str, str], canonical_headers: List[str]) -> Dict[str, str]:
    normalized = {}
    valid_targets = set(canonical_headers)
    for orig, target in mapping.items():
        if target in valid_targets:
            normalized[orig] = target
        else:
            normalized[orig] = orig
    return normalized


def llm_map_headers(source_headers: List[str], target_headers: List[str], model: str = "qwen-plus") -> Dict[str, str]:
    """Use the project's LLM to semantically map source headers to target headers.

    Returns a mapping {source: target} for confident mappings only.
    Falls back to empty mapping on any failure.
    """
    try:
        # Lazy import agent's _invoke_model and ExtractionSettings to call LLM
        from agent import _invoke_model, ExtractionSettings
    except Exception:
        return {}

    prompt = (
        "Map each source column header to the most semantically matching target header.\n"
        f"Source headers: {source_headers}\n"
        f"Target headers: {target_headers}\n\n"
        "Rules:\n"
        "- \"姓名\" / \"有效证件姓名\" / \"申请人姓名\" all mean the same thing.\n"
        "- \"出生日期\" / \"出生年月\" / \"生日\" all mean the same thing.\n"
        "- Only map when you are confident about the match.\n\n"
        "Return a JSON object mapping source->target, e.g. {\"valid_id_name\": \"姓名\", ...}.\n"
        "Only include mappings you are confident about."
    )

    try:
        settings = ExtractionSettings(model=model)
        resp = _invoke_model(model, prompt, settings)
        import json
        try:
            parsed = json.loads(resp)
            if isinstance(parsed, dict):
                # Ensure keys/values are strings
                return {str(k): str(v) for k, v in parsed.items()}
        except Exception:
            # Try to extract JSON object from any surrounding text
            import re
            m = re.search(r"\{[\s\S]*\}", resp)
            if m:
                try:
                    parsed = json.loads(m.group())
                    if isinstance(parsed, dict):
                        return {str(k): str(v) for k, v in parsed.items()}
                except Exception:
                    return {}
    except Exception:
        return {}
    return {}


def align_columns(df: pd.DataFrame, reference_headers: List[str], model: str = "local") -> pd.DataFrame:
    """Align columns with local header-normalisation and similarity rules only.

    This function intentionally does not call llm_map_headers. Excel merging
    therefore requires no cloud API key and does not alter Ollama settings.
    """
    current_headers = [str(c) for c in df.columns]
    if current_headers == reference_headers:
        return df[reference_headers]

    suggested = suggest_column_mapping(current_headers, reference_headers)
    rename_map = {
        source: target
        for source, target in suggested.items()
        if source != target and target in reference_headers
    }
    renamed = df.rename(columns=rename_map)

    for column in reference_headers:
        if column not in renamed.columns:
            renamed[column] = ""

    return renamed[reference_headers]


def merge_excel_dataframes(
    dataframes: List[pd.DataFrame],
    mappings: List[Dict[str, str]],
    canonical_headers: List[str],
    drop_duplicates: bool = False,
) -> pd.DataFrame:
    """Merge multiple dataframes into a single dataframe aligned to canonical_headers.

    Each dataframe may already contain a '来源文件' column; if not, it should be
    supplied by the caller. The function will align columns semantically and
    concatenate rows.
    """
    normalized = []
    for df, mapping in zip(dataframes, mappings):
        # Apply explicit mapping provided by UI (mapping maps original->canonical)
        rename_map = {col: mapping.get(col, col) for col in df.columns}
        df_renamed = df.rename(columns=rename_map)

        # If the dataframe has a single column and that column contains the whole
        # row (common with some exported files), attempt no further splitting here;
        # rely on align_columns to map/safeguard. Add 来源文件 if missing.
        if "来源文件" not in df_renamed.columns:
            df_renamed["来源文件"] = ""

        # Align to canonical headers (ensure 来源文件 included)
        canonical_with_source = list(canonical_headers)
        if "来源文件" not in canonical_with_source:
            canonical_with_source.append("来源文件")

        df_aligned = align_columns(df_renamed, canonical_with_source)
        normalized.append(df_aligned)

    merged = pd.concat(normalized, ignore_index=True, sort=False)
    if drop_duplicates:
        merged = merged.drop_duplicates(ignore_index=True)
    return merged


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, str):
        text = value.strip()
        return text
    return str(value)


def _is_blank_row(row: pd.Series) -> bool:
    return all(_normalize_cell_value(v) == "" for v in row)


def _clean_dataframe_for_merge(df: pd.DataFrame) -> pd.DataFrame:
    """Clean a dataframe before merging.

    Steps:
    - Assume df was read with header=0 (first row = header).
    - If the first data row duplicates the header, drop that single row.
    - Drop columns starting with 'Unnamed' and drop column named '序号' if present.
    - Drop rows that are entirely empty using dropna(how='all').
    - Normalize cell values to strings and strip whitespace.
    """
    if df is None:
        return pd.DataFrame()

    cleaned = df.copy()
    # Ensure column names are strings
    cleaned.columns = [str(c) for c in cleaned.columns]

    # If the first data row duplicates the header row, drop only that row
    if len(cleaned) >= 1:
        try:
            first_row_vals = ["" if pd.isna(x) else str(x).strip() for x in cleaned.iloc[0].tolist()]
            header_vals = [str(h).strip() for h in cleaned.columns.tolist()]
            if first_row_vals == header_vals:
                cleaned = cleaned.iloc[1:].reset_index(drop=True)
        except Exception:
            # be conservative on errors — do not drop anything
            pass

    # Drop unnamed columns and the serial number column
    cleaned = cleaned.loc[:, ~cleaned.columns.str.startswith("Unnamed", na=False)]
    if "序号" in cleaned.columns:
        cleaned = cleaned.drop(columns=["序号"])

    # Drop rows that are entirely empty (NaN in all columns)
    cleaned = cleaned.dropna(how="all")

    # Normalize cell values to strings and strip whitespace; replace NaN with '' first
    cleaned = cleaned.fillna("").apply(lambda col: col.map(_normalize_cell_value))

    return cleaned.reset_index(drop=True)


def merge_excel_folder(folder_path: str, output_path: Optional[str] = None) -> Dict[str, Any]:
    """Merge all .xlsx files in a folder using a union of all columns.

    Rules:
    - Read the first worksheet of each file only.
    - Read all columns as strings with header=0.
    - Drop unnamed columns and the serial number column.
    - Keep all rows; do not deduplicate.
    - Insert a new first column named 序号 with sequential numbering.
    - Save the result as 汇总结果.xlsx in the same folder by default.
    """
    if not os.path.isdir(folder_path):
        raise FileNotFoundError(f"目标文件夹不存在：{folder_path}")

    if output_path is None:
        output_path = os.path.join(folder_path, "汇总结果.xlsx")

    input_files = []
    for filename in sorted(os.listdir(folder_path)):
        if not filename.lower().endswith(".xlsx"):
            continue
        if filename.lower().startswith("~$"):
            continue
        if os.path.abspath(os.path.join(folder_path, filename)) == os.path.abspath(output_path):
            continue
        input_files.append(os.path.join(folder_path, filename))

    if not input_files:
        raise FileNotFoundError(f"文件夹中没有可合并的 .xlsx 文件：{folder_path}")

    cleaned_dfs: List[pd.DataFrame] = []
    all_columns: List[str] = []

    for file_path in input_files:
        # Load file using smart header detection so files with title+number rows
        # above the real header can still be parsed correctly.
        with open(file_path, "rb") as f:
            file_bytes = f.read()
        df = _read_excel_smart(file_bytes, sheet_name=0)

        # Step 1: Drop ALL columns whose name starts with "Unnamed"
        df = df.loc[:, ~df.columns.str.startswith("Unnamed", na=False)]

        # Step 2: Keep only the FIRST occurrence of any duplicate column name
        df = df.loc[:, ~df.columns.duplicated(keep='first')]

        # Step 3: Drop the 序号 column if it exists
        if "序号" in df.columns:
            df = df.drop(columns=["序号"])

        # Step 4: If the first data row duplicates the header, drop that single row
        if not df.empty and len(df) >= 1:
            try:
                first_row_vals = ["" if pd.isna(x) else str(x).strip() for x in df.iloc[0].values]
                header_vals = [str(h).strip() for h in df.columns.tolist()]
                if first_row_vals == header_vals:
                    df = df.iloc[1:].reset_index(drop=True)
            except Exception:
                pass

        # Step 5: Drop fully empty rows
        df = df.dropna(how='all')

        # Normalize cell values to strings and strip whitespace
        df = df.fillna("").apply(lambda col: col.map(lambda v: str(v).strip()))

        if df.empty:
            continue

        cleaned_dfs.append(df)
        all_columns = list(dict.fromkeys(all_columns + list(df.columns)))

    if not cleaned_dfs:
        raise ValueError("所有 Excel 文件都为空，无法生成合并结果")

    aligned_dfs = []
    for df in cleaned_dfs:
        aligned = df.reindex(columns=all_columns, fill_value="")
        aligned_dfs.append(aligned)

    merged = pd.concat(aligned_dfs, ignore_index=True, sort=False)
    merged.insert(0, "序号", range(1, len(merged) + 1))
    merged.to_excel(output_path, index=False)

    # Auto-adjust column widths, capped at 40 characters.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]
        for idx, col in enumerate(merged.columns):
            values = [str(v) for v in merged[col].fillna("").tolist()]
            max_len = max(len(str(col)), *(len(v) for v in values if v is not None)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 40)

    return {
        "success": True,
        "output_path": output_path,
        "files_merged": len(cleaned_dfs),
        "total_rows": len(merged),
        "columns": list(merged.columns),
    }


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Merged") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns):
            values = [str(v) for v in df[col].fillna("").tolist()]
            max_len = max(len(str(col)), *(len(v) for v in values if v is not None)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 40)
    buf.seek(0)
    return buf.getvalue()


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge all .xlsx files in a folder")
    parser.add_argument("folder", help="Folder containing .xlsx files to merge")
    args = parser.parse_args()
    result = merge_excel_folder(args.folder)
    print(result)


if __name__ == "__main__":
    main()


@dataclass
class MergePreview:
    filename: str
    sheet_name: str
    columns: List[str]
    mapping: Dict[str, str]


def preview_merge(files: List[Tuple[str, bytes, str]]) -> Tuple[List[MergePreview], List[str], List[str], float, bool]:
    previews: List[MergePreview] = []
    errors: List[str] = []
    headers_list: List[List[str]] = []
    for filename, file_bytes, sheet_name in files:
        try:
            columns = read_excel_sheet_columns(file_bytes, sheet_name, filename)
        except Exception as exc:
            errors.append(str(exc))
            continue
        headers_list.append(columns)

    if errors:
        return previews, errors, [], 0.0, False

    canonical = build_canonical_headers(headers_list)
    mappings = [suggest_column_mapping(headers, canonical) for headers in headers_list]
    exact = all_headers_exact_match(headers_list)
    score = compute_similarity_score(headers_list)

    for (filename, _, sheet_name), columns, mapping in zip(files, headers_list, mappings):
        previews.append(MergePreview(filename=filename, sheet_name=sheet_name, columns=columns, mapping=mapping))

    return previews, [], canonical, score, exact
