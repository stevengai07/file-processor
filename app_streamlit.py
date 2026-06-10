# -*- coding: utf-8 -*-
import streamlit as st
import tempfile, os, io, json, csv, re
from datetime import datetime
from dotenv import load_dotenv, set_key
from agent import run_agent, build_agent, AVAILABLE_MODELS, PROVIDER_ENV_KEYS
from image_enhancer import PRESETS
from translations import TRANSLATIONS
from extractor import extract_text

load_dotenv()

st.set_page_config(
    page_title="AI Document Workspace",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    [data-testid="stSidebarNav"] {display: none;}
    .result-card {
        background: #1e1e2e; border: 1px solid #313244;
        border-radius: 10px; padding: 16px 20px; margin-bottom: 12px;
    }
    .result-card h4 {
        color: #cba6f7; margin: 0 0 8px 0;
        font-size: 0.85rem; text-transform: uppercase; letter-spacing: 0.08em;
    }
    .result-card p, .result-card li { color: #cdd6f4; font-size: 0.95rem; margin: 4px 0; }
    .tag { display: inline-block; background: #313244; color: #89dceb;
           border-radius: 999px; padding: 2px 12px; margin: 3px 4px 3px 0; font-size: 0.82rem; }
    .sentiment-positive { color: #a6e3a1 !important; font-weight: 600; }
    .sentiment-negative { color: #f38ba8 !important; font-weight: 600; }
    .sentiment-neutral  { color: #fab387 !important; font-weight: 600; }
    .stat-box { text-align: center; background: #1e1e2e; border: 1px solid #313244;
                border-radius: 10px; padding: 14px 8px; }
    .stat-box .num   { font-size: 1.8rem; font-weight: 700; color: #cba6f7; }
    .stat-box .label { font-size: 0.78rem; color: #6c7086; margin-top: 2px; }
    .console-msg-user {
        background: #313244; border-left: 3px solid #cba6f7;
        border-radius: 8px; padding: 10px 14px; margin: 6px 0;
        color: #cdd6f4; font-size: 0.92rem;
    }
    .console-msg-ai {
        background: #1e1e2e; border-left: 3px solid #a6e3a1;
        border-radius: 8px; padding: 10px 14px; margin: 6px 0;
        color: #cdd6f4; font-size: 0.92rem; white-space: pre-wrap;
    }
    .console-msg-system {
        background: #181825; border-left: 3px solid #fab387;
        border-radius: 8px; padding: 8px 14px; margin: 4px 0;
        color: #6c7086; font-size: 0.82rem; font-style: italic;
    }
    .console-wrap {
        background: #11111b; border: 1px solid #313244;
        border-radius: 12px; padding: 16px; max-height: 520px;
        overflow-y: auto; margin-bottom: 12px;
    }
    .tab-header {
        font-size: 0.8rem; color: #6c7086;
        text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 6px;
    }
    .excel-ready-bar {
        background: #1e2e1e; border: 1px solid #3a5a3a;
        border-radius: 10px; padding: 12px 16px; margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
defaults = {
    "lang": "English",
    "result": None,
    "run_count": 0,
    "history": [],
    "last_file": "unknown",
    "last_model": "",
    "extracted_text": "",
    "console_messages": [],
    "console_excel_queue": [],   # list of (label, bytes) ready to download
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

def T(key):
    return TRANSLATIONS.get(st.session_state.lang, TRANSLATIONS["English"]).get(key, key)

# ── Language name → instruction string ────────────────────────────────────────
LANG_INSTRUCTIONS = {
    "English":           "Respond in English.",
    "中文 (Chinese)":    "请用简体中文回复。",
    "Français (French)": "Réponds en français.",
    "Español (Spanish)": "Responde en español.",
    "日本語 (Japanese)": "日本語で回答してください。",
    "한국어 (Korean)":   "한국어로 답변해 주세요.",
}

# ── Quick commands ─────────────────────────────────────────────────────────────
def get_quick_commands():
    return {
        T("qc_summarize"):    "Give me a concise 3-paragraph summary of this document.",
        T("qc_translate_en"): "Translate the full document into English.",
        T("qc_translate_zh"): "Translate the full document into Chinese (Simplified).",
        T("qc_key_facts"):    "List all key facts, statistics, and figures mentioned in the document.",
        T("qc_actions"):      "Extract all action items, tasks, and deadlines from the document.",
        T("qc_qa"):           "I will ask questions about this document. Start by confirming you have read it.",
        T("qc_grammar"):      "Fix all grammar, spelling, and punctuation errors and return the corrected text.",
        T("qc_tables"):       "Extract all tables and structured data from the document in a clean format.",
        T("qc_pros_cons"):    "List the pros and cons / advantages and disadvantages mentioned or implied.",
        T("qc_email"):        "Draft a professional email summarizing the key points of this document.",
        T("qc_categorize"):   "Categorize and tag this document: topic, type, department, urgency level.",
        T("qc_sensitive"):    "Identify any sensitive, confidential, or PII in this document.",
        T("qc_timeline"):     "Extract and organize all dates and events into a chronological timeline.",
        T("qc_insights"):     "What are the 5 most important insights or takeaways from this document?",
    }

# ── Markdown table → Excel bytes ──────────────────────────────────────────────
def markdown_table_to_excel(md_text: str, sheet_title: str = "AI Output") -> bytes | None:
    """
    Finds ALL markdown tables in md_text and writes each as a separate sheet.
    Returns None if no table is found.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill("solid", fgColor="4B2E83")
    BODY_FONT = Font(name="Calibri", size=11)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center", indent=1)
    THIN      = Side(style="thin", color="DDDDDD")
    BRD       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    lines = md_text.splitlines()
    table_blocks = []
    current = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            current.append(stripped)
        else:
            if current:
                table_blocks.append(current)
                current = []
    if current:
        table_blocks.append(current)

    if not table_blocks:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    for t_idx, block in enumerate(table_blocks):
        # Filter out separator rows (---|---|---)
        rows = [row for row in block if not re.match(r"^\|[\s\-:|]+\|$", row)]
        if len(rows) < 2:
            continue

        parsed = []
        for row in rows:
            cells = [c.strip() for c in row.strip("|").split("|")]
            # Strip markdown bold (**text**)
            cells = [re.sub(r"\*\*(.*?)\*\*", r"\1", c) for c in cells]
            parsed.append(cells)

        headers = parsed[0]
        data_rows = parsed[1:]
        ncols = len(headers)

        ws_name = f"Table {t_idx+1}" if len(table_blocks) > 1 else sheet_title[:31]
        ws = wb.create_sheet(title=ws_name)
        ws.column_dimensions["A"].width = 3
        ws.row_dimensions[1].height = 8

        # Title row
        ws.merge_cells(f"B2:{chr(66 + ncols)}2")
        ws["B2"].value = sheet_title
        ws["B2"].font  = Font(name="Calibri", bold=True, size=14, color="4B2E83")
        ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 8

        # Header row
        for ci, h in enumerate(headers):
            col = ci + 2
            c = ws.cell(row=4, column=col, value=h)
            c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BRD
            ws.column_dimensions[chr(64+col)].width = max(16, len(h)+4)
        ws.row_dimensions[4].height = 22

        # Data rows
        for ri, row_data in enumerate(data_rows):
            excel_row = ri + 5
            for ci, val in enumerate(row_data[:ncols]):
                col = ci + 2
                c = ws.cell(row=excel_row, column=col, value=val)
                c.font=BODY_FONT; c.alignment=LEFT; c.border=BRD
                cur_w = ws.column_dimensions[chr(64+col)].width
                ws.column_dimensions[chr(64+col)].width = min(60, max(cur_w, len(val)+4))
            ws.row_dimensions[excel_row].height = 20

        # Excel Table object
        if data_rows:
            last_row = 4 + len(data_rows)
            last_col_letter = chr(64 + ncols + 1)
            tbl = Table(
                displayName=f"AITable{t_idx+1}",
                ref=f"B4:{last_col_letter}{last_row}"
            )
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4", showRowStripes=True
            )
            ws.add_table(tbl)

        ws.freeze_panes = "B5"

    if not wb.sheetnames:
        return None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def response_has_table(text: str) -> bool:
    lines = text.splitlines()
    table_lines = [l for l in lines if l.strip().startswith("|") and l.strip().endswith("|")]
    return len(table_lines) >= 3


def is_excel_intent(user_msg: str) -> bool:
    keywords = [
        "excel", "xlsx", "spreadsheet", "表格", "电子表格", "导出", "生成excel",
        "tableau", "feuille", "hoja", "スプレッドシート", "엑셀", "比较表", "对比表",
        "compare", "comparison", "export", "download", "table"
    ]
    return any(kw in user_msg.lower() for kw in keywords)


# ── Excel builder (extraction results) ────────────────────────────────────────
@st.cache_data(show_spinner=False)
def build_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HDR_FILL   = PatternFill("solid", fgColor="4B2E83")
    TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2E2E2E")
    SUB_FONT   = Font(name="Calibri", italic=True, size=10, color="6C6C6C")
    BODY_FONT  = Font(name="Calibri", size=11)
    LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="4B2E83")
    SENT_POS   = Font(name="Calibri", bold=True, size=11, color="1E8449")
    SENT_NEG   = Font(name="Calibri", bold=True, size=11, color="C0392B")
    SENT_NEU   = Font(name="Calibri", bold=True, size=11, color="D68910")
    TAG_FONT   = Font(name="Calibri", size=10, color="1A73E8")
    WRAP   = Alignment(wrap_text=True, vertical="top", indent=1)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left", vertical="center", indent=1)
    THIN   = Side(style="thin", color="DDDDDD")
    BRD    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    MFILL  = PatternFill("solid", fgColor="F5F0FF")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 72
    ws1.row_dimensions[1].height = 8
    ws1.merge_cells("B2:C2")
    ws1["B2"].value = "AI Document Extraction Report"
    ws1["B2"].font  = TITLE_FONT
    ws1["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 32
    ws1.merge_cells("B3:C3")
    ws1["B3"].value = f"Generated: {data['extracted_at'][:16]}  |  Model: {data['model']}  |  File: {data['file']}"
    ws1["B3"].font  = SUB_FONT
    ws1["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[3].height = 18
    ws1.row_dimensions[4].height = 8

    fields = [
        ("Title",        data["title"],                                           26),
        ("Summary",      data["summary"],                                         52),
        ("Key Topics",   " | ".join(data["key_topics"]),                          26),
        ("Entities",     " | ".join(data["entities"]) or "None found",            26),
        ("Dates",        " | ".join(data["dates"]) if data["dates"] else "None",  26),
        ("Action Items", "\n".join(data["action_items"]) or "None found",         52),
        ("Sentiment",    data["sentiment"],                                        26),
    ]
    for i, (label, value, rh) in enumerate(fields, start=5):
        ws1.row_dimensions[i].height = rh
        lc = ws1.cell(row=i, column=2, value=label)
        lc.font=LABEL_FONT; lc.fill=MFILL; lc.alignment=LEFT; lc.border=BRD
        vc = ws1.cell(row=i, column=3, value=value)
        vc.font=BODY_FONT; vc.alignment=WRAP; vc.border=BRD
        if label == "Sentiment":
            s = (value or "").lower()
            vc.font = SENT_POS if "pos" in s else (SENT_NEG if "neg" in s else SENT_NEU)
        if label == "Key Topics":
            vc.font = TAG_FONT

    ws2 = wb.create_sheet("Topics & Entities")
    ws2.column_dimensions["A"].width=3; ws2.column_dimensions["B"].width=20; ws2.column_dimensions["C"].width=45
    ws2.row_dimensions[1].height=8
    ws2.merge_cells("B2:C2")
    ws2["B2"].value = "Topics & Entities"
    ws2["B2"].font  = Font(name="Calibri", bold=True, size=14, color="4B2E83")
    ws2.row_dimensions[2].height=26; ws2.row_dimensions[3].height=8
    for col, txt in [(2,"Type"),(3,"Value")]:
        c = ws2.cell(row=4, column=col, value=txt)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BRD
    ws2.row_dimensions[4].height=22
    te = [("Key Topic",t) for t in data["key_topics"]] + [("Entity",e) for e in data["entities"]]
    for i,(typ,val) in enumerate(te, start=5):
        fill = PatternFill("solid",fgColor="EDE7F6") if typ=="Key Topic" else PatternFill("solid",fgColor="E8F5E9")
        for col,v in [(2,typ),(3,val)]:
            c = ws2.cell(row=i, column=col, value=v)
            c.font=BODY_FONT; c.fill=fill; c.alignment=LEFT; c.border=BRD
        ws2.row_dimensions[i].height=20
    if te:
        tbl2 = Table(displayName="TopicsEntities", ref=f"B4:C{4+len(te)}")
        tbl2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws2.add_table(tbl2)

    ws3 = wb.create_sheet("Action Items")
    ws3.column_dimensions["A"].width=3; ws3.column_dimensions["B"].width=8
    ws3.column_dimensions["C"].width=65; ws3.column_dimensions["D"].width=18
    ws3.row_dimensions[1].height=8
    ws3.merge_cells("B2:D2")
    ws3["B2"].value = "Action Items"
    ws3["B2"].font  = Font(name="Calibri", bold=True, size=14, color="4B2E83")
    ws3.row_dimensions[2].height=26; ws3.row_dimensions[3].height=8
    for col,txt in [(2,"#"),(3,"Action Item"),(4,"Status")]:
        c = ws3.cell(row=4, column=col, value=txt)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BRD
    ws3.row_dimensions[4].height=22
    actions = data["action_items"] or []
    for i,action in enumerate(actions, start=5):
        rf = PatternFill("solid", fgColor="FAFAFA")
        for col,val in [(2,i-4),(3,action),(4,"Pending")]:
            c = ws3.cell(row=i, column=col, value=val)
            c.font=BODY_FONT; c.fill=rf; c.alignment=(WRAP if col==3 else CENTER); c.border=BRD
        ws3.row_dimensions[i].height=28
    if not actions:
        c = ws3.cell(row=5, column=3, value="No action items found.")
        c.font=Font(name="Calibri", italic=True, color="999999"); c.alignment=LEFT
    if actions:
        tbl3 = Table(displayName="ActionItems", ref=f"B4:D{4+len(actions)}")
        tbl3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium5", showRowStripes=True)
        ws3.add_table(tbl3)

    ws4 = wb.create_sheet("Raw JSON")
    ws4.column_dimensions["A"].width=3; ws4.column_dimensions["B"].width=100
    ws4.row_dimensions[1].height=8
    ws4["B2"].value = "Raw JSON Output"
    ws4["B2"].font  = Font(name="Calibri", bold=True, size=14, color="4B2E83")
    ws4.row_dimensions[2].height=26
    for i,line in enumerate(json.dumps(data, indent=2, ensure_ascii=False).split("\n"), start=4):
        c = ws4.cell(row=i, column=2, value=line)
        c.font=Font(name="Courier New", size=10); c.alignment=LEFT

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ── Console AI call ────────────────────────────────────────────────────────────
def console_ask(user_msg: str, document_text: str, model: str) -> str:
    from langchain_core.messages import SystemMessage, HumanMessage

    lang_instruction = LANG_INSTRUCTIONS.get(st.session_state.lang, "Respond in English.")

    history_context = ""
    for m in st.session_state.console_messages[-6:]:
        role = "User" if m["role"] == "user" else "Assistant"
        history_context += f"{role}: {m['content']}\n"

    table_hint = ""
    if is_excel_intent(user_msg):
        table_hint = (
            "\n\nIMPORTANT: The user wants tabular/Excel output. "
            "You MUST format ALL data as one or more markdown tables using | column | column | syntax. "
            "Do NOT skip the table and just describe it. Output the full table."
        )

    system_prompt = (
        f"{lang_instruction} "
        "You are a highly capable AI document assistant. "
        "You have been given a document to work with. "
        "You can perform ANY task the user asks including: "
        "summarizing, translating, editing, rewriting, extracting data, answering questions, "
        "creating tables, drafting emails, finding inconsistencies, checking grammar, "
        "building timelines, and more. "
        "Always be thorough, precise, and helpful. "
        "Format your response clearly using markdown when appropriate."
        + table_hint
    )

    doc_context = f"\n\n---DOCUMENT---\n{document_text[:10000]}\n---END DOCUMENT---" if document_text else ""
    history_ctx = f"\n\n---CONVERSATION HISTORY---\n{history_context}---END HISTORY---" if history_context else ""

    provider = AVAILABLE_MODELS[model][0]
    api_key  = os.getenv(PROVIDER_ENV_KEYS[provider])

    if provider == "OpenAI":
        from langchain_openai import ChatOpenAI
        llm = ChatOpenAI(model=model, temperature=0.3, openai_api_key=api_key)
    elif provider == "Anthropic":
        from langchain_anthropic import ChatAnthropic
        llm = ChatAnthropic(model=model, temperature=0.3, anthropic_api_key=api_key)
    elif provider == "Google":
        from langchain_google_genai import ChatGoogleGenerativeAI
        llm = ChatGoogleGenerativeAI(model=model, temperature=0.3, google_api_key=api_key)
    elif provider == "DeepSeek":
        from langchain_openai import ChatOpenAI
        llm = ChatOpenAI(model=model, temperature=0.3, openai_api_key=api_key,
                         base_url="https://api.deepseek.com/v1")
    elif provider == "xAI":
        from langchain_openai import ChatOpenAI
        llm = ChatOpenAI(model=model, temperature=0.3, openai_api_key=api_key,
                         base_url="https://api.x.ai/v1")
    elif provider == "Alibaba":
        from langchain_openai import ChatOpenAI
        llm = ChatOpenAI(model=model, temperature=0.3, openai_api_key=api_key,
                         base_url="https://dashscope.aliyuncs.com/compatible-mode/v1")
    else:
        raise ValueError(f"Unknown provider: {provider}")

    messages = [
        SystemMessage(content=system_prompt + doc_context + history_ctx),
        HumanMessage(content=user_msg),
    ]
    return llm.invoke(messages).content


# ── Process a console message and handle Excel output ─────────────────────────
def process_console_message(user_msg: str):
    with st.spinner("..."):
        try:
            reply = console_ask(user_msg, st.session_state.extracted_text, model_choice)
            st.session_state.console_messages.append({"role": "assistant", "content": reply})

            # Auto-generate Excel if response contains a table
            if response_has_table(reply):
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                label = f"ai_output_{ts}.xlsx"
                xl_bytes = markdown_table_to_excel(reply, sheet_title="AI Output")
                if xl_bytes:
                    st.session_state.console_excel_queue.append((label, xl_bytes))
        except Exception as e:
            st.session_state.console_messages.append(
                {"role": "assistant", "content": f"❌ Error: {e}"}
            )


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"## {T('settings')}")
    st.divider()

    lang_options = list(TRANSLATIONS.keys())
    st.session_state.lang = st.selectbox(
        T("language"), lang_options,
        index=lang_options.index(st.session_state.lang)
    )
    st.divider()

    st.markdown(f"### {T('model_section')}")
    providers = sorted(set(v[0] for v in AVAILABLE_MODELS.values()))
    provider_choice = st.selectbox(T("provider"), providers)
    provider_models = {k: v for k, v in AVAILABLE_MODELS.items() if v[0] == provider_choice}
    model_choice = st.selectbox(T("model"), list(provider_models.keys()))
    _, cost, desc = AVAILABLE_MODELS[model_choice]
    st.caption(f"💰 {cost}  •  {desc}")
    st.divider()

    st.markdown(f"### {T('scanned_docs')}")
    preset_map = {
        "scanner":      T("preset_scanner"),
        "phone_camera": T("preset_phone"),
        "faded":        T("preset_faded"),
        "fax":          T("preset_fax"),
    }
    preset_choice = st.selectbox(
        T("enhancement_preset"), list(PRESETS.keys()),
        format_func=lambda x: preset_map.get(x, x)
    )
    ocr_lang = st.selectbox(T("ocr_language"), ["eng","chi_sim","fra","deu","jpn","kor"])
    st.divider()

    st.markdown(f"### {T('api_key_section')}")
    env_key = PROVIDER_ENV_KEYS[provider_choice]
    current_val = os.getenv(env_key, "")
    masked = ("*" * 8 + current_val[-4:]) if len(current_val) > 4 else ""
    status = T("api_key_set") if current_val else T("api_key_not_set")
    st.caption(f"{'🟢' if current_val else '🔴'} `{env_key}` {status}")
    if masked:
        st.caption(f"{T('api_key_current')}: `{masked}`")
    new_key = st.text_input(T("update_api_key"), type="password", placeholder=T("paste_new_key"))
    if st.button(T("save_key"), use_container_width=True):
        if new_key.strip():
            env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
            set_key(env_path, env_key, new_key.strip())
            os.environ[env_key] = new_key.strip()
            st.success(T("key_saved"))
        else:
            st.warning(T("key_empty"))
    st.divider()

    st.markdown(f"### {T('session_stats')}")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(
            f'<div class="stat-box">'
            f'<div class="num">{st.session_state.run_count}</div>'
            f'<div class="label">{T("extractions")}</div>'
            f'</div>',
            unsafe_allow_html=True)
    with c2:
        st.markdown(
            f'<div class="stat-box">'
            f'<div class="num">{len(st.session_state.history)}</div>'
            f'<div class="label">{T("history")}</div>'
            f'</div>',
            unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(T("app_title"))
st.caption(T("app_caption"))

tab_extract, tab_console = st.tabs([T("tab_extract"), T("tab_console")])


# ──────────────────────────────────────────────────────────────────────────────
# TAB 1: Extract & Export
# ──────────────────────────────────────────────────────────────────────────────
with tab_extract:
    uploaded = st.file_uploader(T("drop_file"), type=["pdf","docx"], label_visibility="visible")

    col_run, col_clear, col_hist = st.columns([2,1,1])
    with col_run:
        run_btn = st.button(f"⚡ {T('extract_button')}", use_container_width=True,
                            type="primary", disabled=(uploaded is None))
    with col_clear:
        if st.button(T("clear"), use_container_width=True):
            st.session_state.result = None
            st.rerun()
    with col_hist:
        show_history = st.toggle(T("history_toggle"), value=False)

    if run_btn and uploaded:
        suffix = ".pdf" if uploaded.name.lower().endswith(".pdf") else ".docx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded.read())
            tmp_path = tmp.name
        with st.spinner(f"{T('extracting')} {model_choice}..."):
            try:
                raw_text = extract_text(tmp_path, ocr_preset=preset_choice, lang=ocr_lang)
                st.session_state.extracted_text = raw_text
                result = run_agent(tmp_path, ocr_preset=preset_choice, lang=ocr_lang, model=model_choice)
                st.session_state.result     = result
                st.session_state.run_count += 1
                st.session_state.last_file  = uploaded.name
                st.session_state.last_model = model_choice
                st.session_state.history.append({"file": uploaded.name, "model": model_choice, "result": result})
                st.session_state.console_messages.append({
                    "role": "system",
                    "content": (
                        f"📄 {T('doc_loaded')}: {uploaded.name} "
                        f"({len(raw_text.split()):,} {T('words')}). "
                        f"{T('tab_console')} tab is ready."
                    )
                })
                st.success(
                    f"{T('done_with')} **{model_choice}** — "
                    f"switch to {T('tab_console')} to interact with this document."
                )
            except Exception as e:
                st.error(f"{T('error')}: {e}")
            finally:
                os.unlink(tmp_path)

    if show_history and st.session_state.history:
        st.markdown("---")
        st.markdown(T("history_title"))
        for i, h in enumerate(reversed(st.session_state.history)):
            with st.expander(f"#{len(st.session_state.history)-i}  {h['file']}  •  {h['model']}"):
                st.write(f"**{T('title_label')}:** {h['result'].title}")
                st.write(f"**{T('summary_label')}:** {h['result'].summary}")

    if st.session_state.result:
        r = st.session_state.result
        st.markdown("---")
        st.markdown(T("extracted_info"))

        col1, col2 = st.columns([3,1])
        with col1:
            st.markdown(
                f'<div class="result-card"><h4>{T("title_card")}</h4>'
                f'<p style="font-size:1.1rem;font-weight:600;">{r.title}</p></div>',
                unsafe_allow_html=True)
        with col2:
            sent = (r.sentiment or "neutral").lower()
            sc   = "sentiment-positive" if "pos" in sent else ("sentiment-negative" if "neg" in sent else "sentiment-neutral")
            icon = "😊" if "pos" in sent else ("😟" if "neg" in sent else "😐")
            st.markdown(
                f'<div class="result-card" style="text-align:center;"><h4>{T("sentiment_card")}</h4>'
                f'<p class="{sc}" style="font-size:1.3rem;">{icon} {r.sentiment}</p></div>',
                unsafe_allow_html=True)

        st.markdown(
            f'<div class="result-card"><h4>{T("summary_card")}</h4><p>{r.summary}</p></div>',
            unsafe_allow_html=True)

        col3, col4 = st.columns(2)
        with col3:
            tags = "".join(f'<span class="tag">🏷️ {t}</span>' for t in r.key_topics)
            st.markdown(
                f'<div class="result-card"><h4>{T("topics_card")}</h4><div>{tags}</div></div>',
                unsafe_allow_html=True)
        with col4:
            etags = "".join(f'<span class="tag">👤 {e}</span>' for e in r.entities)
            entities_content = etags if etags else f"<p>{T('none_found')}</p>"
            st.markdown(
                f'<div class="result-card"><h4>{T("entities_card")}</h4>'
                f'<div>{entities_content}</div></div>',
                unsafe_allow_html=True)

        col5, col6 = st.columns(2)
        with col5:
            dhtml = "".join(f"<li>📅 {d}</li>" for d in r.dates) if r.dates else f"<p>{T('none_found')}</p>"
            st.markdown(
                f'<div class="result-card"><h4>{T("dates_card")}</h4>'
                f'<ul style="margin:0;padding-left:18px;">{dhtml}</ul></div>',
                unsafe_allow_html=True)
        with col6:
            ahtml = "".join(f"<li>✅ {a}</li>" for a in r.action_items) if r.action_items else f"<p>{T('none_found')}</p>"
            st.markdown(
                f'<div class="result-card"><h4>{T("actions_card")}</h4>'
                f'<ul style="margin:0;padding-left:18px;">{ahtml}</ul></div>',
                unsafe_allow_html=True)

        st.markdown("---")
        st.markdown(T("export_title"))
        result_dict = {
            "extracted_at": datetime.now().isoformat(),
            "model":        st.session_state.last_model,
            "file":         st.session_state.last_file,
            "title":        r.title,
            "summary":      r.summary,
            "key_topics":   r.key_topics,
            "entities":     r.entities,
            "dates":        r.dates,
            "action_items": r.action_items,
            "sentiment":    r.sentiment,
        }
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dcol1, dcol2, dcol3 = st.columns(3)
        with dcol1:
            st.download_button(T("download_json"),
                data=json.dumps(result_dict, indent=2, ensure_ascii=False),
                file_name=f"extraction_{ts}.json", mime="application/json",
                use_container_width=True)
        with dcol2:
            csv_buf = io.StringIO()
            w = csv.writer(csv_buf)
            w.writerow(["Field","Value"])
            for field, val in [
                ("Title",        r.title),
                ("Summary",      r.summary),
                ("Key Topics",   " | ".join(r.key_topics)),
                ("Entities",     " | ".join(r.entities)),
                ("Dates",        " | ".join(r.dates) if r.dates else ""),
                ("Action Items", " | ".join(r.action_items) if r.action_items else ""),
                ("Sentiment",    r.sentiment),
                ("Model",        st.session_state.last_model),
                ("Extracted At", result_dict["extracted_at"]),
            ]:
                w.writerow([field, val])
            st.download_button(T("download_csv"), data=csv_buf.getvalue(),
                file_name=f"extraction_{ts}.csv", mime="text/csv",
                use_container_width=True)
        with dcol3:
            st.download_button(T("download_excel"), data=build_excel(result_dict),
                file_name=f"extraction_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 2: AI Console
# ──────────────────────────────────────────────────────────────────────────────
with tab_console:
    has_doc = bool(st.session_state.extracted_text)

    hcol1, hcol2, hcol3 = st.columns([3,1,1])
    with hcol1:
        if has_doc:
            word_count = len(st.session_state.extracted_text.split())
            st.caption(f"📄 {T('doc_loaded')}: **{st.session_state.last_file}** — {word_count:,} {T('words')}")
        else:
            st.caption(f"⚠️ {T('no_doc_warning')}")
    with hcol2:
        if st.button(T("clear_chat"), use_container_width=True):
            st.session_state.console_messages = []
            st.session_state.console_excel_queue = []
            st.rerun()
    with hcol3:
        if st.session_state.console_messages:
            transcript = "\n\n".join(
                f"[{m['role'].upper()}]\n{m['content']}"
                for m in st.session_state.console_messages
            )
            st.download_button(T("transcript"), data=transcript,
                file_name=f"console_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain", use_container_width=True)

    # ── Excel download queue (auto-generated from AI responses) ───────────────
    if st.session_state.console_excel_queue:
        st.markdown("---")
        st.markdown("#### 📊 Excel Files Ready")
        for idx, (fname, xl_bytes) in enumerate(st.session_state.console_excel_queue):
            ecol1, ecol2 = st.columns([4,1])
            with ecol1:
                st.caption(f"📥 `{fname}`")
            with ecol2:
                st.download_button(
                    T("download_excel"),
                    data=xl_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"xl_dl_{idx}"
                )

    st.markdown("---")

    # ── Quick command buttons
    st.markdown(f'<p class="tab-header">{T("quick_commands")}</p>', unsafe_allow_html=True)
    QUICK_COMMANDS = get_quick_commands()
    qcols = st.columns(7)
    for i, (label, prompt) in enumerate(QUICK_COMMANDS.items()):
        with qcols[i % 7]:
            if st.button(label, use_container_width=True, disabled=not has_doc, key=f"qc_{i}"):
                st.session_state.console_messages.append({"role": "user", "content": prompt})
                process_console_message(prompt)
                st.rerun()

    st.markdown("---")

    # ── Message history
    if st.session_state.console_messages:
        msgs_html = ""
        for m in st.session_state.console_messages:
            content = m["content"].replace("<","&lt;").replace(">","&gt;").replace("\n","<br>")
            if m["role"] == "user":
                msgs_html += f'<div class="console-msg-user">👤 <strong>You:</strong><br>{content}</div>'
            elif m["role"] == "assistant":
                msgs_html += f'<div class="console-msg-ai">🤖 <strong>AI:</strong><br>{content}</div>'
            else:
                msgs_html += f'<div class="console-msg-system">ℹ️ {content}</div>'
        st.markdown(f'<div class="console-wrap">{msgs_html}</div>', unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="console-wrap">'
            f'<div class="console-msg-system">ℹ️ {T("console_empty")}</div>'
            f'</div>',
            unsafe_allow_html=True)

    # ── Input form
    with st.form("console_form", clear_on_submit=True):
        user_input = st.text_area(
            "...", placeholder=T("input_placeholder"),
            height=90, label_visibility="collapsed"
        )
        scol1, scol2 = st.columns([5,1])
        with scol1:
            submitted = st.form_submit_button(T("send"), use_container_width=True,
                                               type="primary", disabled=not has_doc)
        with scol2:
            st.form_submit_button(T("cancel"), use_container_width=True)

    if submitted and user_input.strip():
        st.session_state.console_messages.append({"role": "user", "content": user_input.strip()})
        process_console_message(user_input.strip())
        st.rerun()