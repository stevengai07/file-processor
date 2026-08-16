import io
import zipfile

from openpyxl import Workbook, load_workbook

from order_matching import (
    assign_unique_matches,
    build_matching_report,
    build_renamed_zip,
    calculate_match_score,
    confirm_match,
    extract_document_identity,
    make_output_filename,
    normalise_code,
    normalise_title,
    parse_order_excel,
)
from schema import DocumentCandidate, MatchStatus, OrderItem, OrderingTask


def _workbook_bytes(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "顺序"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_order_excel_auto_mapping():
    raw = _workbook_bytes(["发言顺序", "材料名称", "材料编号"], [[1, "标题一", "A-001"]])
    parsed = parse_order_excel(raw, "order.xlsx")
    assert not parsed.errors
    assert parsed.items[0].order_no == 1
    assert parsed.items[0].title == "标题一"
    assert parsed.items[0].code == "A-001"


def test_parse_order_excel_duplicate_order():
    raw = _workbook_bytes(["顺序", "标题"], [[1, "标题一"], [1, "标题二"]])
    parsed = parse_order_excel(raw, "order.xlsx")
    assert any("重复" in e for e in parsed.errors)


def test_parse_order_excel_empty_title_and_code_invalid():
    raw = _workbook_bytes(["顺序", "标题", "编号"], [[1, "", ""]])
    parsed = parse_order_excel(raw, "order.xlsx")
    assert any("标题和编号" in e for e in parsed.errors)


def test_parse_order_excel_agenda_without_header():
    raw = _workbook_bytes([], [])
    wb = Workbook()
    ws = wb.active
    ws.append([1, "9:00-9:08", "低碳技术创新功能型平台", "上海"])
    ws.append([2, "9:08-9:16", "智能型新能源汽车研发与转化功能型平台", "上海"])
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_order_excel(buf.getvalue(), "agenda.xlsx")

    assert not parsed.errors
    assert parsed.column_mapping["order_no"] == "__col_1__"
    assert parsed.items[0].order_no == 1
    assert parsed.items[0].time_range == "9:00-9:08"
    assert parsed.items[0].title == "低碳技术创新功能型平台"
    assert parsed.items[0].region == "上海"
    assert parsed.items[0].code is None


def test_parse_order_excel_real_agenda_with_intro_and_group_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["附件", None, None, None])
    ws.append(["研发载体发言顺序表", None, None, None])
    ws.append(["4月23日上午", None, None, None])
    ws.append(["序号", "时间", "研发载体名称", "所在地"])
    ws.append([1, "9:00-9:08", "低碳技术创新功能型平台", "上海"])
    ws.append([2, "9:08-9:16", "智能型新能源汽车研发与转化功能型平台", "上海"])
    ws.append(["4月23日下午", None, None, None])
    ws.append([24, "13:30-13:38", "微技术工业研究院", "上海"])
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_order_excel(buf.getvalue(), "agenda.xlsx")

    assert not parsed.errors
    assert parsed.auto_header_row == 4
    assert parsed.header_row == 4
    assert parsed.skipped_rows == 1
    assert len(parsed.items) == 3
    assert parsed.items[0].title == "低碳技术创新功能型平台"
    assert parsed.items[0].time_range == "9:00-9:08"
    assert parsed.items[0].region == "上海"
    assert parsed.items[2].order_no == 24
    assert parsed.items[2].title == "微技术工业研究院"


def test_code_normalisation_and_match():
    assert normalise_code("ａ-００１ 号") == "A001号".upper()
    item = OrderItem(item_id="i1", order_no=1, title="", code="A-001", source_row=2)
    doc = DocumentCandidate(file_id="f1", filename="x.pdf", extension=".pdf", extracted_code="a 001")
    score, method, _ = calculate_match_score(item, doc)
    assert score == 0.98
    assert method == "code_normalised"


def test_title_matching_modes():
    assert normalise_title("《关于年度预算执行情况的报告材料》.pdf") == "关于年度预算执行情况的报告"
    item = OrderItem(item_id="i1", order_no=1, title="关于年度预算执行情况的报告", code=None, source_row=2)
    doc = DocumentCandidate(file_id="f1", filename="x.pdf", extension=".pdf", extracted_title="关于年度预算执行情况的报告")
    assert calculate_match_score(item, doc)[0] == 0.95
    doc.extracted_title = "关于年度预算执行情况的报告补充说明"
    assert calculate_match_score(item, doc)[0] == 0.90


def test_assign_unique_matches_prevents_duplicate_auto_assignment():
    items = [
        OrderItem(item_id="i1", order_no=1, title="标题", code="A001", source_row=2),
        OrderItem(item_id="i2", order_no=2, title="标题", code="A001", source_row=3),
    ]
    docs = [DocumentCandidate(file_id="f1", filename="a.pdf", extension=".pdf", extracted_code="A001")]
    matches = assign_unique_matches(items, docs)
    assert matches[0].selected_file_id == "f1"
    assert matches[1].status == MatchStatus.NEEDS_REVIEW


def test_manual_confirmation_and_reassignment():
    match = assign_unique_matches(
        [OrderItem(item_id="i1", order_no=1, title="标题", code=None, source_row=2)],
        [DocumentCandidate(file_id="f1", filename="a.pdf", extension=".pdf", extracted_title="标题")],
    )[0]
    confirm_match(match, "f1")
    assert match.status == MatchStatus.CONFIRMED
    assert match.manually_confirmed is True


def test_output_filename_sanitises_and_deduplicates():
    used = set()
    item = OrderItem(item_id="i1", order_no=1, title='标题 /:*?"<>|', code=None, source_row=2)
    first = make_output_filename(item, ".pdf", used)
    second = make_output_filename(item, ".pdf", used)
    assert first == "001_标题 ________.pdf"
    assert second.endswith("_2.pdf")


def test_zip_contains_chinese_filename_and_report():
    item = OrderItem(item_id="i1", order_no=1, title="关于预算的报告", code="A001", source_row=2)
    doc = DocumentCandidate(file_id="f1", filename="原文件.pdf", extension=".pdf", extracted_title="关于预算的报告", extracted_code="A001")
    task = OrderingTask(ordering_id="t1", order_filename="order.xlsx", items=[item], documents=[doc])
    task.matches = assign_unique_matches(task.items, task.documents)
    zip_bytes = build_renamed_zip(task, {"f1": b"PDFDATA"})
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        assert "001_关于预算的报告.pdf" in names
        assert "匹配清单.xlsx" in names
        assert zf.read("001_关于预算的报告.pdf") == b"PDFDATA"


def test_matching_report_content():
    item = OrderItem(item_id="i1", order_no=1, title="标题", code="A001", source_row=2)
    doc = DocumentCandidate(file_id="f1", filename="原文件.pdf", extension=".pdf", extracted_title="标题", extracted_code="A001")
    task = OrderingTask(ordering_id="t1", order_filename="order.xlsx", items=[item], documents=[doc])
    task.matches = assign_unique_matches(task.items, task.documents)
    report = build_matching_report(task)
    wb = load_workbook(io.BytesIO(report))
    ws = wb.active
    assert ws[1][0].value == "顺序号"
    assert ws[2][0].value == 1
    assert ws[2][5].value == "原文件.pdf"


def test_ppt_uses_filename_identity_without_body_extraction():
    candidate = extract_document_identity("001_低碳技术创新功能型平台.ppt", b"not-real-ppt", use_ollama=False)
    assert candidate.extension == ".ppt"
    assert candidate.extracted_title == "低碳技术创新功能型平台"
    assert "文件名" in candidate.extraction_error
